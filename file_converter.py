import xlrd
import tkinter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from datetime import datetime
from tkinter import filedialog, StringVar, messagebox, ttk, Label
from ttkthemes import ThemedTk


class Converter:
    def __init__(self, master):
        self.master = master
        self.path = StringVar()
        self.message = StringVar()
        self.__main_window()
        self.__style()
        self.__form()

    def __main_window(self):
        self.master.title("Eclispe Timesheet Converter")
        tkinter.Tk.iconbitmap(self.master, default="converter.ico")
        self.master.resizable(0, 0)

    def __style(self):
        self.style = ttk.Style()
        self.style.configure("Text", padding=6, anchor="left")
        self.style.configure("TEntry", bg="white")

    def __form(self):
        self.f_label = Label(
            self.master, text="Find Source File:", fg="dimgrey")
        self.f_label.grid(row=1, column=1, padx=10, pady=10)
        self.f_entry = ttk.Entry(
            self.master, textvariable=self.path, width=40, style="TEntry")
        self.f_entry.grid(row=1, column=2, padx=10, pady=10)
        self.f_button = ttk.Button(
            self.master, text="Browse", command=self.__browse_button)
        self.f_button.grid(row=1, column=3, padx=10, pady=10)
        self.f_help = Label(
            self.master, fg="dimgrey",
            text=(
                "Please select the .xls file downloaded from Eclipse you want "
                "to convert."))
        self.f_help.grid(row=2, columnspan=6, padx=10)
        self.createBut = ttk.Button(
            self.master, text="Convert File", command=self.__convert)
        self.createBut.grid(row=5, columnspan=4, padx=10, pady=10)

    def __browse_button(self):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xls")])
        self.path.set(filename)
        self.path_name = str(filename)
        folder_loc = filename.rsplit('/', 1)[0]
        self.src_name = str(folder_loc)

    def __convert_to_xlsx(self):
        "Loads the new xlsx file and makes the current worksheet active"
        self.filename = self.path_name.rsplit('/', 1)[1].replace('.xls', '')
        book_xls = xlrd.open_workbook(self.path.get())
        book_xlsx = Workbook()
        s_names = book_xls.sheet_names()
        for sheet in range(0, len(s_names)):
            sheet_xls = book_xls.sheet_by_name(s_names[sheet])
            if sheet == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = s_names[sheet]
            else:
                sheet_xlsx = book_xlsx.create_sheet(
                    title=s_names[sheet])
            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(
                        row=row+1, column=col+1).value = sheet_xls.cell_value(
                            row, col)
        book_xlsx.save(f'{self.src_name}/{self.filename}.xlsx')
        self.xlsx_file = f'{self.src_name}/{self.filename}.xlsx'

    def __load_xlsx(self):
        "Loads the new xlsx file and makes the current worksheet active"
        self.workbook = load_workbook(self.xlsx_file)
        self.worksheet = self.workbook.active

    def __remove_rows(self):
        "Removes lines from the worksheet that are not needed."
        ws = self.worksheet
        # remove footer rows
        ws.delete_rows(ws.max_row - 1, ws.max_row)
        # remove blank header row
        ws.delete_rows(1, 1)
        for cell in ws['A'][1:]:
            # remove header rows from other pages
            if cell.value == 'Expenses':
                ws.delete_rows(cell.row)
        for cell in ws['I'][1:]:
            # remove candidate total lines
            if cell.value == 'VAT:':
                ws.delete_rows(cell.row, 1)
        self.workbook.save(self.xlsx_file)

    def __candidate_data(self):
        ws = self.worksheet
        self.candidates = {}
        data_store = {}
        can_id = None
        can_len = 0
        for cell in ws['L'][1:]:
            row_id = str(cell.row)
            if cell.value is None and ws[f'E{row_id}'].value == 'Sort code:':
                can_id = ws[f'A{row_id}'].value
                name = ws[f'B{row_id}'].value
                can_name = self.__check_name_length(name)
                can_len = 0
                self.candidates[can_id] = {
                    can_len: {
                        'id': can_id,
                        'first': can_name[0],
                        'surname': can_name[1]}}
                data_store = {
                    'id': can_id,
                    'first': can_name[0],
                    'surname': can_name[1]}
            elif cell.value is not None:
                date = self.__date_convert(ws[f'B{row_id}'].value)
                ts_data = {
                    'ts_id': ws[f'A{row_id}'].value,
                    'pe_date': date,
                    'rate_desc': ws[f'D{row_id}'].value,
                    'rate_hours': float(ws[f'F{row_id}'].value),
                    'rate': float(ws[f'E{row_id}'].value)}
                self.candidates[can_id][can_len] = {**data_store, **ts_data}
                data_store.update({
                    'ts_id': ws[f'A{row_id}'].value,
                    'pe_date': date,
                    'rate_desc': '',
                    'rate_hours': '',
                    'rate': ''})
                can_len += 1
            else:
                other_data = {
                    'rate_desc': ws[f'A{row_id}'].value,
                    'rate_hours': float(ws[f'C{row_id}'].value),
                    'rate': float(ws[f'B{row_id}'].value)}
                self.candidates[can_id][can_len] = {**data_store, **other_data}
                can_len += 1

    def __check_name_length(self, val):
        name_part = val.split()
        name = [name_part[0], name_part[1]]
        if len(name_part) > 2:
            for part in range(2, len(name_part)):
                name[1] = name[1] + ' ' + name_part[part]
        return name

    def __date_convert(self, date):
        dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + date - 2)
        return dt.date().strftime('%d/%m/%Y')

    def __data_row(self, sheet):
        r_index = 1
        for x in self.candidates:
            for i in self.candidates[x]:
                r_index += 1
                sheet[f'A{r_index}'] = self.candidates[x][i]['ts_id']
                sheet[f'C{r_index}'] = self.candidates[x][i]['id']
                sheet[f'D{r_index}'] = self.candidates[x][i]['first']
                sheet[f'E{r_index}'] = self.candidates[x][i]['surname']
                sheet[f'F{r_index}'] = 'HOURS'
                sheet[f'G{r_index}'] = self.candidates[x][i]['rate']
                sheet[f'H{r_index}'] = self.candidates[x][i]['rate_hours']
                sheet[f'J{r_index}'] = self.candidates[x][i]['rate_desc']

    def __create_output(self):
        header_list = [
            'Contract', 'ID', 'Reference', 'Name', 'Surname', 'Rate Type',
            'Rate', 'Unit', 'Company ID', 'Description', 'Cost Code',
            'Rate Deduction', 'Tax Rate %']
        # Open new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.filename}_converted"
        # Add header row
        for row in range(1, 2):
            for col in range(1, len(header_list)+1):
                ws.cell(column=col, row=row, value=header_list[col-1])
        # Add data lines
        self.__data_row(ws)
        # Save workbook to same folder as source file
        wb.save(f'{self.src_name}/{self.filename}_converted.xlsx')

    def __convert(self):
        self.__convert_to_xlsx()
        self.__load_xlsx()
        self.__remove_rows()
        self.__candidate_data()
        self.__create_output()
        messagebox.showinfo("Process Complete", "The app will now close")
        self.master.destroy()


def main():
    root = ThemedTk(theme="arc")
    app = Converter(root)
    root.mainloop()


if __name__ == '__main__':
    main()
